"""
admin/semesters.py
Quản lý Semester và danh sách Contest trong mỗi semester.
"""
import datetime
import re

from flask import flash, jsonify, redirect, render_template, request, url_for

from CTFd.admin import admin
from CTFd.models import Users, db
from CTFd.models import Contests as Contest, ContestParticipants as ContestParticipant, Semester
from CTFd.utils.decorators import admins_only
from CTFd.plugins import bypass_csrf_protection


# ── helpers ───────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "-", text)
    return text[:100]


def _unique_slug(base: str, exclude_id: int | None = None) -> str:
    slug = _slugify(base)
    q = Contest.query.filter_by(slug=slug)
    if exclude_id:
        q = q.filter(Contest.id != exclude_id)
    if q.first():
        slug = f"{slug}-{int(datetime.datetime.utcnow().timestamp())}"
    return slug


# ── Semester CRUD ─────────────────────────────────────────────────────────────

@admin.route("/admin/semesters")
@admins_only
def semesters_listing():
    semesters = Semester.query.order_by(Semester.id.desc()).all()
    return render_template("admin/semesters/listing.html", semesters=semesters)


@admin.route("/admin/semesters/new", methods=["GET", "POST"])
@admins_only
def semester_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Tên kỳ học không được trống.", "danger")
            return redirect(url_for("admin.semester_new"))

        if Semester.query.filter_by(semester_name=name).first():
            flash(f"Kỳ học '{name}' đã tồn tại.", "danger")
            return redirect(url_for("admin.semester_new"))

        def _parse_dt(field):
            val = request.form.get(field, "").strip()
            if not val:
                return None
            try:
                return datetime.datetime.fromisoformat(val)
            except ValueError:
                return None

        sem = Semester(
            semester_name=name,
            start_time=_parse_dt("start_time"),
            end_time=_parse_dt("end_time"),
        )
        db.session.add(sem)
        db.session.commit()
        flash(f"Đã tạo kỳ học '{sem.semester_name}'.", "success")
        return redirect(url_for("admin.semester_detail", semester_id=sem.id))

    return render_template("admin/semesters/new.html")


@admin.route("/admin/semesters/<int:semester_id>")
@admins_only
def semester_detail(semester_id):
    sem = Semester.query.get_or_404(semester_id)
    contests = (
        Contest.query.filter_by(semester_name=sem.semester_name)
        .order_by(Contest.created_at.desc())
        .all()
    )
    for c in contests:
        c.participant_count = ContestParticipant.query.filter_by(contest_id=c.id).count()
    return render_template(
        "admin/semesters/detail.html", semester=sem, contests=contests
    )


@admin.route("/admin/semesters/<int:semester_id>/edit", methods=["GET", "POST"])
@admins_only
def semester_edit(semester_id):
    sem = Semester.query.get_or_404(semester_id)
    if request.method == "POST":
        new_name = request.form.get("name", sem.semester_name).strip()
        if not new_name:
            flash("Tên kỳ học không được trống.", "danger")
            return redirect(url_for("admin.semester_edit", semester_id=sem.id))

        conflict = Semester.query.filter(
            Semester.semester_name == new_name,
            Semester.id != sem.id
        ).first()
        if conflict:
            flash(f"Tên '{new_name}' đã được dùng bởi kỳ học khác.", "danger")
            return redirect(url_for("admin.semester_edit", semester_id=sem.id))

        def _parse_dt(field, fallback=None):
            val = request.form.get(field, "").strip()
            if not val:
                return fallback
            try:
                return datetime.datetime.fromisoformat(val)
            except ValueError:
                return fallback

        old_name = sem.semester_name
        if new_name != old_name:
            Contest.query.filter_by(semester_name=old_name).update(
                {"semester_name": new_name}, synchronize_session=False
            )

        sem.semester_name = new_name
        sem.start_time = _parse_dt("start_time", sem.start_time)
        sem.end_time = _parse_dt("end_time", sem.end_time)
        db.session.commit()
        flash("Đã cập nhật kỳ học.", "success")
        return redirect(url_for("admin.semester_detail", semester_id=sem.id))
    return render_template("admin/semesters/edit.html", semester=sem)


@admin.route("/admin/semesters/<int:semester_id>/delete", methods=["POST"])
@admins_only
def semester_delete(semester_id):
    sem = Semester.query.get_or_404(semester_id)
    name = sem.semester_name
    db.session.delete(sem)
    db.session.commit()
    flash(f"Đã xoá kỳ học '{name}'.", "success")
    return redirect(url_for("admin.semesters_listing"))


# ── Contest CRUD (trong một semester) ────────────────────────────────────────

@admin.route("/admin/semesters/<int:semester_id>/contests/new", methods=["GET", "POST"])
@admins_only
def contest_new(semester_id):
    sem = Semester.query.get_or_404(semester_id)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Tên contest không được trống.", "danger")
            return redirect(url_for("admin.contest_new", semester_id=semester_id))

        slug = _unique_slug(request.form.get("slug", "") or name)

        def _parse_dt(field):
            val = request.form.get(field, "").strip()
            if not val:
                return None
            try:
                return datetime.datetime.fromisoformat(val)
            except ValueError:
                return None

        from CTFd.utils.user import get_current_user
        owner = get_current_user()

        contest = Contest(
            semester_name=sem.semester_name,
            name=name,
            description=request.form.get("description", "").strip() or None,
            slug=slug,
            owner_id=owner.id if owner else None,
            state=request.form.get("state", "hidden"),
            user_mode=request.form.get("user_mode", "users"),
            start_time=_parse_dt("start_time"),
            end_time=_parse_dt("end_time"),
            freeze_scoreboard_at=_parse_dt("freeze_scoreboard_at"),
        )
        db.session.add(contest)
        db.session.commit()
        flash(f"Đã tạo contest '{contest.name}'.", "success")
        return redirect(
            url_for("admin.contest_dashboard", contest_id=contest.id)
        )

    return render_template("admin/semesters/contest_new.html", semester=sem)


@admin.route("/admin/contests/<int:contest_id>/edit", methods=["GET", "POST"])
@admins_only
def contest_edit(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    if request.method == "POST":
        contest.name = request.form.get("name", contest.name).strip()
        contest.description = request.form.get("description", "").strip() or None
        contest.state = request.form.get("state", contest.state)
        contest.user_mode = request.form.get("user_mode", contest.user_mode)
        contest.updated_at = datetime.datetime.utcnow()

        def _parse_dt(field, fallback=None):
            val = request.form.get(field, "").strip()
            if not val:
                return fallback
            try:
                return datetime.datetime.fromisoformat(val)
            except ValueError:
                return fallback

        contest.start_time = _parse_dt("start_time", contest.start_time)
        contest.end_time = _parse_dt("end_time", contest.end_time)
        contest.freeze_scoreboard_at = _parse_dt(
            "freeze_scoreboard_at", contest.freeze_scoreboard_at
        )

        # Slug chỉ thay nếu người dùng nhập mới
        new_slug = request.form.get("slug", "").strip()
        if new_slug and new_slug != contest.slug:
            contest.slug = _unique_slug(new_slug, exclude_id=contest.id)

        db.session.commit()
        flash("Đã cập nhật contest.", "success")
        return redirect(url_for("admin.contest_dashboard", contest_id=contest.id))

    return render_template("admin/contests/edit.html", contest=contest)


@admin.route("/admin/contests/<int:contest_id>/delete", methods=["POST"])
@admins_only
def contest_delete(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    semester_name = contest.semester_name
    name = contest.name
    db.session.delete(contest)
    db.session.commit()
    flash(f"Đã xoá contest '{name}'.", "success")
    if semester_name:
        return redirect(url_for("admin.semesters_listing"))
    return redirect(url_for("admin.semesters_listing"))


# ── Contest Dashboard (landing khi vào một contest) ───────────────────────────

@admin.route("/admin/contests/<int:contest_id>")
@admins_only
def contest_dashboard(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    participant_count = ContestParticipant.query.filter_by(contest_id=contest_id).count()
    from CTFd.models import ContestsChallenges as ContestChallenge, Submissions, Solves, Fails, db as _db
    from CTFd.utils.modes import get_model

    challenge_count = ContestChallenge.query.filter_by(contest_id=contest_id).count()

    cc_ids = [
        cc.id for cc in ContestChallenge.query.filter_by(contest_id=contest_id).with_entities(ContestChallenge.id).all()
    ]
    solve_count = Solves.query.filter(Solves.contest_challenge_id.in_(cc_ids)).count() if cc_ids else 0
    submission_count = Submissions.query.filter(Submissions.contest_challenge_id.in_(cc_ids)).count() if cc_ids else 0
    wrong_count = Fails.query.filter(Fails.contest_challenge_id.in_(cc_ids)).count() if cc_ids else 0

    # Total possible points
    total_points = _db.session.query(_db.func.sum(ContestChallenge.value)).filter_by(contest_id=contest_id).scalar() or 0

    # Solve counts per challenge for most/least solved
    from CTFd.models import Challenges
    solves_per_cc = {}
    if cc_ids:
        rows = (
            _db.session.query(Solves.contest_challenge_id, _db.func.count(Solves.id).label("cnt"))
            .filter(Solves.contest_challenge_id.in_(cc_ids))
            .group_by(Solves.contest_challenge_id)
            .all()
        )
        cc_map = {cc.id: cc for cc in ContestChallenge.query.filter_by(contest_id=contest_id).all()}
        for cc_id, cnt in rows:
            cc = cc_map.get(cc_id)
            name = cc.name if cc and cc.name else (Challenges.query.get(cc.bank_id).name if cc and cc.bank_id else f"#{cc_id}")
            solves_per_cc[name] = cnt

    most_solved = max(solves_per_cc, key=solves_per_cc.get) if solves_per_cc else None
    least_solved = min(solves_per_cc, key=solves_per_cc.get) if solves_per_cc else None

    return render_template(
        "admin/contests/dashboard.html",
        contest=contest,
        participant_count=participant_count,
        challenge_count=challenge_count,
        solve_count=solve_count,
        submission_count=submission_count,
        wrong_count=wrong_count,
        total_points=total_points,
        solve_data=solves_per_cc,
        most_solved=most_solved,
        least_solved=least_solved,
    )


# ── Contest → Challenges ──────────────────────────────────────────────────────

@admin.route("/admin/contests/<int:contest_id>/challenges")
@admins_only
def contest_challenges(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    from CTFd.models import ContestsChallenges as ContestChallenge, Challenges as ChallengeBank, Challenges, Tags

    # Lấy danh sách challenge_id thuộc contest này
    cc_bank_ids = {
        cc.bank_id: cc
        for cc in ContestChallenge.query.filter_by(contest_id=contest_id).all()
    }
    bank_challenge_ids = list(cc_bank_ids.keys())

    # Map bank_id → challenge_id trong Challenges table (nếu có liên kết)
    # Thực tế contest dùng Challenges model (legacy), filter theo bank_id
    # Lấy Challenges có id nằm trong bank_challenge_ids
    q = request.args.get("q")
    field = request.args.get("field") or "name"
    category = request.args.get("category")
    type_ = request.args.get("type")
    difficulty = request.args.get("difficulty")
    state_filter = request.args.get("state")
    has_prereq = request.args.get("has_prereq")
    page = abs(request.args.get("page", 1, type=int))
    tags_q = request.args.get("tags")

    tag_terms = []
    filters = [Challenges.id.in_(bank_challenge_ids)] if bank_challenge_ids else [db.false()]

    if tags_q:
        tag_terms = [t.strip() for t in tags_q.split(",") if t.strip()]
        for term in tag_terms:
            exists_filter = (
                db.session.query(Tags.id)
                .filter(Tags.challenge_id == Challenges.id, db.func.lower(Tags.value) == term.lower())
                .exists()
            )
            filters.append(exists_filter)

    if q and Challenges.__mapper__.has_property(field):
        filters.append(getattr(Challenges, field).like(f"%{q}%"))
    if category:
        filters.append(Challenges.category == category)
    if type_:
        filters.append(Challenges.type == type_)
    if difficulty:
        filters.append(Challenges.difficulty == int(difficulty))
    if state_filter:
        filters.append(Challenges.state == state_filter)
    if has_prereq == "yes":
        filters.append(Challenges.requirements.isnot(None))
    elif has_prereq == "no":
        filters.append(Challenges.requirements.is_(None))

    challenges = Challenges.query.filter(*filters).order_by(Challenges.id.asc()).paginate(
        page=page, per_page=50, error_out=False
    )

    # Thêm creator name
    for c in challenges.items:
        c.creator = c.author.name if c.author else "Unknown"

    categories = [r[0] for r in db.session.query(Challenges.category).filter(
        Challenges.id.in_(bank_challenge_ids)).distinct().all() if r[0]]
    types = [r[0] for r in db.session.query(Challenges.type).filter(
        Challenges.id.in_(bank_challenge_ids)).distinct().all() if r[0]]

    args = dict(request.args)
    args.pop("page", None)
    args["contest_id"] = contest_id

    prev_page = url_for("admin.contest_challenges", page=challenges.prev_num, **args) if challenges.has_prev else "#"
    next_page = url_for("admin.contest_challenges", page=challenges.next_num, **args) if challenges.has_next else "#"

    return render_template(
        "admin/contests/challenges.html",
        contest=contest,
        challenges=challenges,
        field=field,
        q=q,
        category=category,
        type=type_,
        difficulty=difficulty,
        state_filter=state_filter,
        has_prereq=has_prereq,
        tag_terms=tag_terms,
        categories=categories,
        types=types,
        prev_page=prev_page,
        next_page=next_page,
    )


@admin.route("/admin/contests/<int:contest_id>/challenges/add", methods=["POST"])
@admins_only
@bypass_csrf_protection
def contest_challenge_add(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    from CTFd.models import ContestsChallenges as ContestChallenge, Challenges as ChallengeBank
    data = request.get_json(silent=True) or {}
    bank_ids = data.get("bank_ids", [])
    if not bank_ids:
        return jsonify({"success": False, "message": "Không có challenge nào được chọn."}), 400

    added, skipped = [], []
    for bid in bank_ids:
        bc = ChallengeBank.query.get(bid)
        if not bc:
            skipped.append(bid)
            continue
        exists = ContestChallenge.query.filter_by(
            contest_id=contest_id, bank_id=bid
        ).first()
        if exists:
            skipped.append(bid)
            continue
        cc = ContestChallenge(
            contest_id=contest_id,
            bank_id=bid,
            name=bc.name,
            connection_protocol=bc.connection_protocol or "http",
            max_deploy_count=bc.max_deploy_count or 0,
        )
        db.session.add(cc)
        added.append(bid)

    db.session.commit()
    return jsonify({
        "success": True,
        "message": f"Đã thêm {len(added)} challenge.",
        "added": added,
        "skipped": skipped,
    })


@admin.route(
    "/admin/contests/<int:contest_id>/challenges/<int:cc_id>/remove",
    methods=["POST"]
)
@admins_only
def contest_challenge_remove(contest_id, cc_id):
    from CTFd.models import ContestsChallenges as ContestChallenge
    cc = ContestChallenge.query.filter_by(
        id=cc_id, contest_id=contest_id
    ).first_or_404()
    db.session.delete(cc)
    db.session.commit()
    flash("Đã xoá challenge khỏi contest.", "success")
    return redirect(url_for("admin.contest_challenges", contest_id=contest_id))


# ── Contest → Users (Participants) ───────────────────────────────────────────

@admin.route("/admin/contests/<int:contest_id>/users")
@admins_only
def contest_users(contest_id):
    contest = Contest.query.get_or_404(contest_id)

    q = request.args.get("q", "").strip()
    field = request.args.get("field") or "name"
    role_filter = request.args.get("role", "")
    verified_filter = request.args.get("verified", "")
    hidden_filter = request.args.get("hidden", "")
    banned_filter = request.args.get("banned", "")
    page = abs(request.args.get("page", 1, type=int))

    # Lấy user_id thuộc contest
    participant_user_ids = [
        cp.user_id for cp in ContestParticipant.query.filter_by(contest_id=contest_id).all()
    ]

    filters = [Users.id.in_(participant_user_ids)] if participant_user_ids else [db.false()]

    if q and hasattr(Users, field):
        filters.append(getattr(Users, field).like(f"%{q}%"))
    if role_filter:
        filters.append(Users.type == role_filter)
    if verified_filter == "true":
        filters.append(Users.verified == True)
    elif verified_filter == "false":
        filters.append(Users.verified == False)
    if hidden_filter == "true":
        filters.append(Users.hidden == True)
    elif hidden_filter == "false":
        filters.append(Users.hidden == False)
    if banned_filter == "true":
        filters.append(Users.banned == True)
    elif banned_filter == "false":
        filters.append(Users.banned == False)

    users_q = Users.query.filter(*filters).order_by(Users.id.asc()).paginate(
        page=page, per_page=50, error_out=False
    )

    args = dict(request.args)
    args.pop("page", None)
    args["contest_id"] = contest_id
    prev_page = url_for("admin.contest_users", page=users_q.prev_num, **args) if users_q.has_prev else "#"
    next_page = url_for("admin.contest_users", page=users_q.next_num, **args) if users_q.has_next else "#"

    return render_template(
        "admin/contests/users.html",
        contest=contest,
        users=users_q,
        q=q,
        field=field,
        role_filter=role_filter,
        verified_filter=verified_filter,
        hidden_filter=hidden_filter,
        banned_filter=banned_filter,
        prev_page=prev_page,
        next_page=next_page,
    )


@admin.route("/admin/contests/<int:contest_id>/users/import", methods=["POST"])
@admins_only
def contest_users_import(contest_id):
    """Import users từ danh sách users tổng vào contest."""
    contest = Contest.query.get_or_404(contest_id)
    data = request.get_json(silent=True) or {}
    user_ids = data.get("user_ids", [])

    if not user_ids:
        return jsonify({"success": False, "message": "Không có user nào được chọn."}), 400

    added, skipped = [], []
    for uid in user_ids:
        user = Users.query.get(uid)
        if not user:
            skipped.append(uid)
            continue
        exists = ContestParticipant.query.filter_by(
            contest_id=contest_id, user_id=uid
        ).first()
        if exists:
            skipped.append(uid)
            continue
        cp = ContestParticipant(
            contest_id=contest_id,
            user_id=uid,
            role="contestant",
            score=0,
        )
        db.session.add(cp)
        added.append(uid)

    db.session.commit()
    return jsonify({
        "success": True,
        "message": f"Đã import {len(added)} user.",
        "added": added,
        "skipped": skipped,
    })


@admin.route(
    "/admin/contests/<int:contest_id>/users/<int:user_id>/remove",
    methods=["POST"]
)
@admins_only
def contest_user_remove(contest_id, user_id):
    cp = ContestParticipant.query.filter_by(
        contest_id=contest_id, user_id=user_id
    ).first_or_404()
    db.session.delete(cp)
    db.session.commit()
    flash("Đã xoá user khỏi contest.", "success")
    return redirect(url_for("admin.contest_users", contest_id=contest_id))


@admin.route("/admin/contests/<int:contest_id>/users/import-excel", methods=["POST"])
@admins_only
@bypass_csrf_protection
def contest_users_import_excel(contest_id):
    """
    Import users vào contest từ file Excel/CSV.
    - File cần có cột 'email'
    - Nếu user đã tồn tại → thêm vào contest_participants
    - Nếu chưa tồn tại → tạo user mới với email đó → thêm vào contest_participants
    """
    import io, secrets, string

    Contest.query.get_or_404(contest_id)

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "message": "Không có file."}), 400

    filename = file.filename.lower()
    emails = []

    try:
        if filename.endswith(".csv"):
            import csv
            content = file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                email = (row.get("email") or row.get("Email") or "").strip().lower()
                if email:
                    emails.append(email)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            headers = [
                str(c.value).strip().lower() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            if "email" not in headers:
                return jsonify({"success": False, "message": "Không tìm thấy cột 'email' trong file."}), 400
            email_col = headers.index("email")
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[email_col]
                if val:
                    emails.append(str(val).strip().lower())
    except Exception as e:
        return jsonify({"success": False, "message": f"Lỗi đọc file: {str(e)}"}), 400

    if not emails:
        return jsonify({"success": False, "message": "Không có email nào trong file."}), 400

    added = skipped = created = 0

    for email in emails:
        user = Users.query.filter_by(email=email).first()

        if not user:
            # Tạo user mới với email, name tạm = phần trước @
            temp_name = email.split("@")[0]
            temp_pass = "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(16))
            user = Users(
                name=temp_name,
                email=email,
                password=temp_pass,
                type="user",
                verified=False,
            )
            db.session.add(user)
            db.session.flush()  # lấy user.id
            created += 1

        exists = ContestParticipant.query.filter_by(
            contest_id=contest_id, user_id=user.id
        ).first()
        if exists:
            skipped += 1
            continue

        db.session.add(ContestParticipant(
            contest_id=contest_id,
            user_id=user.id,
            role="contestant",
            score=0,
        ))
        added += 1

    db.session.commit()
    return jsonify({
        "success": True,
        "added": added,
        "created": created,
        "skipped": skipped,
    })


# ── Contest → Scoreboard ──────────────────────────────────────────────────────

@admin.route("/admin/contests/<int:contest_id>/scoreboard")
@admins_only
def contest_scoreboard(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    standings = (
        ContestParticipant.query
        .filter_by(contest_id=contest_id)
        .order_by(
            ContestParticipant.score.desc(),
            ContestParticipant.last_solve_at.asc(),
        )
        .all()
    )
    # Gắn user
    for p in standings:
        p.user_obj = Users.query.get(p.user_id)
    return render_template(
        "admin/contests/scoreboard.html", contest=contest, standings=standings
    )


# ── API: danh sách users tổng để import ──────────────────────────────────────

@admin.route("/admin/contests/<int:contest_id>/api/available-users")
@admins_only
def contest_api_available_users(contest_id):
    """Trả về users chưa có trong contest (để import)."""
    already_in = db.session.query(ContestParticipant.user_id).filter_by(
        contest_id=contest_id
    ).subquery()
    users = (
        Users.query
        .filter(Users.type == "user")
        .filter(Users.id.not_in(already_in))
        .order_by(Users.name)
        .all()
    )
    return jsonify({
        "success": True,
        "users": [
            {"id": u.id, "name": u.name, "email": u.email}
            for u in users
        ],
    })


# ═══════════════════════════════════════════════════════════════════════════════
# Contest-scoped pages — wrap existing global pages với contest context
# ═══════════════════════════════════════════════════════════════════════════════

@admin.route("/admin/contests/<int:contest_id>/submissions")
@admin.route("/admin/contests/<int:contest_id>/submissions/<type>")
@admins_only
def contest_submissions(contest_id, type=None):
    contest = Contest.query.get_or_404(contest_id)
    from CTFd.models import Submissions, Challenges, Teams as TeamsModel, ContestsChallenges as ContestChallenge

    q = request.args.get("q")
    field = request.args.get("field")
    page = abs(request.args.get("page", 1, type=int))
    team_filter = request.args.get("team_id", "").strip()
    user_filter = request.args.get("user_id", "").strip()
    challenge_filter = request.args.get("challenge_id", "").strip()

    # Lấy contest_challenge_ids thuộc contest này
    cc_ids = [
        cc.id for cc in ContestChallenge.query.filter_by(contest_id=contest_id)
        .with_entities(ContestChallenge.id).all()
    ]

    filters = []
    if cc_ids:
        filters.append(Submissions.contest_challenge_id.in_(cc_ids))
    else:
        filters.append(db.false())
    if type:
        filters.append(Submissions.type == type)
    if team_filter:
        filters.append(Submissions.team_id == int(team_filter))
    if user_filter:
        filters.append(Submissions.user_id == int(user_filter))
    if challenge_filter:
        filters.append(Submissions.contest_challenge_id == int(challenge_filter))

    submissions = (
        Submissions.query.filter(*filters)
        .order_by(Submissions.date.desc())
        .paginate(page=page, per_page=50, error_out=False)
    )

    participant_ids = [cp.user_id for cp in ContestParticipant.query.filter_by(contest_id=contest_id).all()]
    all_users = Users.query.filter(Users.id.in_(participant_ids)).order_by(Users.name).all() if participant_ids else []
    all_teams = TeamsModel.query.filter(TeamsModel.contest_id == contest_id).all()
    all_challenges = ContestChallenge.query.filter_by(contest_id=contest_id).all()

    args = dict(request.args)
    args.pop("page", None)
    prev_page = url_for("admin.contest_submissions", contest_id=contest_id, page=submissions.prev_num, **args) if submissions.has_prev else "#"
    next_page = url_for("admin.contest_submissions", contest_id=contest_id, page=submissions.next_num, **args) if submissions.has_next else "#"

    return render_template(
        "admin/contests/submissions.html",
        contest=contest,
        submissions=submissions,
        type=type,
        q=q,
        field=field,
        team_filter=team_filter,
        user_filter=user_filter,
        challenge_filter=challenge_filter,
        all_teams=all_teams,
        all_users=all_users,
        all_challenges=all_challenges,
        prev_page=prev_page,
        next_page=next_page,
        timezone_offset="",
        filter_args={},
    )


@admin.route("/admin/contests/<int:contest_id>/teams")
@admins_only
def contest_teams(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    from CTFd.models import Teams as TeamsModel

    q = request.args.get("q", "").strip()
    field = request.args.get("field") or "name"
    hidden = request.args.get("hidden", "")
    banned = request.args.get("banned", "")
    page = abs(request.args.get("page", 1, type=int))

    filters = [TeamsModel.contest_id == contest_id]

    if q and hasattr(TeamsModel, field):
        filters.append(getattr(TeamsModel, field).ilike(f"%{q}%"))
    if hidden == "true":
        filters.append(TeamsModel.hidden == True)
    elif hidden == "false":
        filters.append(TeamsModel.hidden == False)
    if banned == "true":
        filters.append(TeamsModel.banned == True)
    elif banned == "false":
        filters.append(TeamsModel.banned == False)

    teams = TeamsModel.query.filter(*filters).order_by(TeamsModel.id.asc()).paginate(
        page=page, per_page=50, error_out=False
    )

    args = dict(request.args)
    args.pop("page", None)
    prev_page = url_for("admin.contest_teams", contest_id=contest_id, page=teams.prev_num, **args) if teams.has_prev else "#"
    next_page = url_for("admin.contest_teams", contest_id=contest_id, page=teams.next_num, **args) if teams.has_next else "#"

    return render_template(
        "admin/contests/teams.html",
        contest=contest,
        teams=teams,
        q=q,
        field=field,
        hidden=hidden,
        banned=banned,
        bracket_id="",
        prev_page=prev_page,
        next_page=next_page,
    )


@admin.route("/admin/contests/<int:contest_id>/teams/new", methods=["GET", "POST"])
@admins_only
def contest_team_new(contest_id):
    contest = Contest.query.get_or_404(contest_id)
    from CTFd.models import Teams as TeamsModel
    from CTFd.utils.crypto import hash_password
    from CTFd.forms.teams import TeamCreateForm

    form = TeamCreateForm()

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        password = (request.form.get("password") or "").strip()
        email = (request.form.get("email") or "").strip() or None

        if not name:
            flash("Tên team không được trống.", "danger")
        elif not password:
            flash("Password không được trống.", "danger")
        elif TeamsModel.query.filter_by(name=name).first():
            flash(f"Team '{name}' đã tồn tại.", "danger")
        else:
            team = TeamsModel(
                name=name,
                email=email,
                password=hash_password(password),
                hidden=request.form.get("hidden") == "y",
                banned=request.form.get("banned") == "y",
                contest_id=contest_id,
            )
            db.session.add(team)
            db.session.commit()
            flash(f"Đã tạo team '{name}'.", "success")
            return redirect(url_for("admin.contest_teams", contest_id=contest_id))

    return render_template(
        "admin/contests/team_new.html",
        contest=contest,
        form=form,
    )


@admin.route("/admin/import_contest_participants", methods=["POST"])
@admins_only
def import_contest_participants():
    from io import StringIO
    import csv
    from CTFd.models import Users, ContestParticipants as ContestParticipant
    from CTFd.utils.crypto import hash_password

    contest_id = request.form.get("contest_id", type=int)
    if not contest_id:
        flash("Invalid contest ID.", "danger")
        return redirect(url_for("admin.semesters_listing"))

    if "csv_file" not in request.files:
        flash("No file provided.", "danger")
        return redirect(url_for("admin.semesters_listing"))

    file = request.files["csv_file"]
    if file.filename == "":
        flash("No file selected.", "danger")
        return redirect(url_for("admin.semesters_listing"))

    try:
        raw = file.stream.read()
        try:
            csvdata = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            csvdata = raw.decode("latin-1")

        csvfile = StringIO(csvdata)
        reader = csv.DictReader(csvfile)

        def normalize_row(row):
            return {k.strip().lower() if isinstance(k, str) else k: v for k, v in row.items()}

        added_users = 0
        added_participants = 0

        for row in reader:
            row = normalize_row(row)
            email = row.get("email")
            if not email:
                continue
            email = email.strip().lower()
            user = Users.query.filter_by(email=email).first()
            if not user:
                name = row.get("name", email.split("@")[0])
                password = row.get("password", "fpt123456")
                user = Users(
                    name=name,
                    email=email,
                    password=hash_password(password),
                    type="user",
                    hidden=False,
                )
                db.session.add(user)
                db.session.commit()
                added_users += 1
            participant = ContestParticipant.query.filter_by(
                contest_id=contest_id, user_id=user.id
            ).first()
            if not participant:
                participant = ContestParticipant(
                    contest_id=contest_id,
                    user_id=user.id,
                )
                db.session.add(participant)
                added_participants += 1

        db.session.commit()
        flash(f"Imported successfully. Created {added_users} new users, added {added_participants} participants to contest.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error importing participants: {str(e)}", "danger")

    return redirect(url_for("admin.semesters_listing"))
